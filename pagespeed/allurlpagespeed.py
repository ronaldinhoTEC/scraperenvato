import requests
import openpyxl
from openpyxl.styles import PatternFill

# Configura tu API Key aquí
API_KEY = 'AIzaSyCJInl7NQUgMHD7sLL7ML57LAjYUTKbqyI'

# Función para obtener las puntuaciones de PageSpeed Insights
def get_pagespeed_scores(url):
    api_url = f"https://www.googleapis.com/pagespeedonline/v5/runPagespeed?url={url}&key={API_KEY}"
    response = requests.get(api_url)
    if response.status_code == 200:
        data = response.json()
        # Extraer los resultados para escritorio y móvil
        desktop_scores = {
            "performance": data['lighthouseResult']['categories']['performance']['score'] * 100,
            "accessibility": data['lighthouseResult']['categories']['accessibility']['score'] * 100,
            "best_practices": data['lighthouseResult']['categories']['best-practices']['score'] * 100,
            "seo": data['lighthouseResult']['categories']['seo']['score'] * 100,
        }
        mobile_scores = {
            "performance": data['lighthouseResult']['categories']['performance']['score'] * 100,
            "accessibility": data['lighthouseResult']['categories']['accessibility']['score'] * 100,
            "best_practices": data['lighthouseResult']['categories']['best-practices']['score'] * 100,
            "seo": data['lighthouseResult']['categories']['seo']['score'] * 100,
        }
        return desktop_scores, mobile_scores
    else:
        print(f"Error al obtener los datos de PageSpeed para {url}: {response.status_code}")
        return None, None

# Función para asignar color basado en la puntuación
def get_fill_color(score):
    if 90 <= score <= 100:
        return PatternFill(start_color="00FF00", end_color="00FF00", fill_type="solid")  # Verde
    elif 50 <= score < 90:
        return PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")  # Amarillo
    else:
        return PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")  # Rojo

# Crear el archivo Excel
def create_excel_report(results):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "PageSpeed Insights"

    # Escribir los encabezados
    headers = ['URL', 'Rendimiento (Escritorio)', 'Accesibilidad (Escritorio)', 'Mejores Prácticas (Escritorio)', 'SEO (Escritorio)',
               'Rendimiento (Móvil)', 'Accesibilidad (Móvil)', 'Mejores Prácticas (Móvil)', 'SEO (Móvil)']
    ws.append(headers)

    # Escribir los resultados en el archivo Excel
    for url, scores in results.items():
        ws.append([url,
                   scores['desktop']['performance'], scores['desktop']['accessibility'], scores['desktop']['best_practices'], scores['desktop']['seo'],
                   scores['mobile']['performance'], scores['mobile']['accessibility'], scores['mobile']['best_practices'], scores['mobile']['seo']])

    # Aplicar los colores correspondientes en las celdas
    for row in ws.iter_rows(min_row=2, min_col=2, max_col=9):
        for cell in row:
            if cell.column == 2:  # Escritorio - Rendimiento
                cell.fill = get_fill_color(cell.value)
            elif cell.column == 3:  # Escritorio - Accesibilidad
                cell.fill = get_fill_color(cell.value)
            elif cell.column == 4:  # Escritorio - Mejores Prácticas
                cell.fill = get_fill_color(cell.value)
            elif cell.column == 5:  # Escritorio - SEO
                cell.fill = get_fill_color(cell.value)
            elif cell.column == 6:  # Móvil - Rendimiento
                cell.fill = get_fill_color(cell.value)
            elif cell.column == 7:  # Móvil - Accesibilidad
                cell.fill = get_fill_color(cell.value)
            elif cell.column == 8:  # Móvil - Mejores Prácticas
                cell.fill = get_fill_color(cell.value)
            elif cell.column == 9:  # Móvil - SEO
                cell.fill = get_fill_color(cell.value)

    # Guardar el archivo Excel
    wb.save("pagespeed_results.xlsx")
    print("El archivo Excel se ha guardado correctamente como 'pagespeed_results.xlsx'.")

# Leer las URLs del archivo "plantillas_completas.txt"
def read_urls():
    with open("plantillas_completas.txt", "r", encoding="utf-8") as file:
        lines = file.readlines()

    urls = []
    for line in lines:
        if 'URL de la plantilla:' in line:
            url = line.split("URL de la plantilla:")[1].strip()
            urls.append(url)
    return urls

# Ejecutar el proceso de análisis y creación de archivo Excel
def main():
    urls = read_urls()
    results = {}

    for url in urls:
        print(f"Evaluando {url}...")
        desktop_scores, mobile_scores = get_pagespeed_scores(url)
        if desktop_scores and mobile_scores:
            results[url] = {
                "desktop": desktop_scores,
                "mobile": mobile_scores,
            }
        else:
            print(f"Error al evaluar {url}.")
    
    create_excel_report(results)

if __name__ == "__main__":
    main()
