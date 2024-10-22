import requests
import openpyxl
from openpyxl.styles import PatternFill, Alignment

# Configura tu API Key aquí
API_KEY = 'AIzaSyCJInl7NQUgMHD7sLL7ML57LAjYUTKbqyI'

# URL de la API de PageSpeed Insights
API_URL = "https://www.googleapis.com/pagespeedonline/v5/runPagespeed"

# Función para obtener los puntajes de PageSpeed Insights
def get_pagespeed_data(url, strategy="desktop"):
    # Parámetros de la solicitud
    params = {
        'url': url,
        'key': API_KEY,
        'category': ['performance', 'accessibility', 'best-practices', 'seo'],
        'strategy': strategy  # Añadir estrategia para escritorio o móvil
    }

    # Realizar la solicitud a la API
    response = requests.get(API_URL, params=params)

    if response.status_code == 200:
        data = response.json()

        # Obtener los puntajes de cada categoría con un valor por defecto si no existen
        performance_score = data.get('lighthouseResult', {}).get('categories', {}).get('performance', {}).get('score', 0) * 100
        accessibility_score = data.get('lighthouseResult', {}).get('categories', {}).get('accessibility', {}).get('score', 0) * 100
        best_practices_score = data.get('lighthouseResult', {}).get('categories', {}).get('best-practices', {}).get('score', 0) * 100
        seo_score = data.get('lighthouseResult', {}).get('categories', {}).get('seo', {}).get('score', 0) * 100

        return performance_score, accessibility_score, best_practices_score, seo_score
    else:
        print(f"Error al acceder a la API para {url}. Código de estado: {response.status_code}")
        return 0, 0, 0, 0  # Valores por defecto

# Función para asignar color a las celdas en función del puntaje
def get_fill_color(score):
    if 90 <= score <= 100:
        return PatternFill(start_color="00FF00", end_color="00FF00", fill_type="solid")  # Verde
    elif 50 <= score < 90:
        return PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")  # Amarillo
    else:
        return PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")  # Rojo

# Crear el archivo Excel con los resultados
def create_excel_report(results):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "PageSpeed Insights"

    # Encabezados para el archivo Excel
    headers = ['URL', 'Rendimiento (Escritorio)', 'Accesibilidad (Escritorio)', 'Mejores Prácticas (Escritorio)', 'SEO (Escritorio)',
               'Rendimiento (Móvil)', 'Accesibilidad (Móvil)', 'Mejores Prácticas (Móvil)', 'SEO (Móvil)']
    ws.append(headers)

    # Escribir los resultados en el archivo Excel
    for url, scores in results.items():
        ws.append([url,
                   scores['desktop']['performance'], scores['desktop']['accessibility'], scores['desktop']['best_practices'], scores['desktop']['seo'],
                   scores['mobile']['performance'], scores['mobile']['accessibility'], scores['mobile']['best_practices'], scores['mobile']['seo']])

    # Aplicar colores según los puntajes
    for row in ws.iter_rows(min_row=2, min_col=2, max_col=9):
        for cell in row:
            cell.fill = get_fill_color(cell.value)

    # Ajustar el ancho de las columnas para una mejor presentación
    column_widths = [15, 25, 25, 25, 25, 25, 25, 25, 25]
    for i, col_width in enumerate(column_widths, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = col_width

    # Guardar el archivo Excel
    wb.save("pagespeed_results.xlsx")
    print("El archivo Excel se ha guardado correctamente como 'pagespeed_results.xlsx'.")

# Leer las URLs desde el archivo
def read_urls():
    with open("urls_limpias.txt", "r", encoding="utf-8") as file:
        urls = [line.strip() for line in file.readlines()]
    return urls

# Ejecutar el proceso de análisis y creación de archivo Excel
def main():
    urls = read_urls()

    # Pedir al usuario el rango de URLs a evaluar
    start = int(input("Introduce el número de inicio del rango de URLs que quieres evaluar: "))
    end = int(input("Introduce el número de fin del rango de URLs que quieres evaluar: "))

    # Validar el rango
    urls = urls[start-1:end]

    results = {}

    for url in urls:
        print(f"Evaluando {url}...")  # Solo mostrar que URL se está evaluando
        desktop_scores = get_pagespeed_data(url, strategy="desktop")
        mobile_scores = get_pagespeed_data(url, strategy="mobile")
        if desktop_scores and mobile_scores:
            results[url] = {
                "desktop": {
                    "performance": desktop_scores[0],
                    "accessibility": desktop_scores[1],
                    "best_practices": desktop_scores[2],
                    "seo": desktop_scores[3],
                },
                "mobile": {
                    "performance": mobile_scores[0],
                    "accessibility": mobile_scores[1],
                    "best_practices": mobile_scores[2],
                    "seo": mobile_scores[3],
                },
            }
        else:
            print(f"Error al evaluar {url}.")

    create_excel_report(results)

if __name__ == "__main__":
    main()
